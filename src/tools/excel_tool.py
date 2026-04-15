"""
엑셀 출력 도구

3시트 구성:
  배당 데이터   : 검증 통과 항목 전체
  수동 확인 필요 : 자동 검증 실패 항목
  검증 로그     : 전체 실행 요약
"""
from __future__ import annotations

import logging
import os
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 컬럼 정의
_MAIN_COLS = [
    ("ticker",           "종목코드"),
    ("company_name",     "종목명"),
    ("year",             "연도"),
    ("dividend_amount",  "주당배당금(원)"),
    ("dividend_yield",   "배당수익률(%)"),
    ("ex_dividend_date", "배당락일"),
    ("record_date",      "배당기준일"),
    ("payment_date",     "배당지급일"),
    ("dividend_status",  "배당확정여부"),
    ("confidence_score", "신뢰도"),
    ("sources",          "데이터출처"),
]

_MANUAL_COLS = _MAIN_COLS + [
    ("validation_reason", "검증실패사유"),
    ("retry_count",       "재시도횟수"),
]

_LOG_COLS = [
    ("ticker",            "종목코드"),
    ("company_name",      "종목명"),
    ("year",              "연도"),
    ("validation_status", "검증상태"),
    ("confidence_score",  "신뢰도"),
    ("retry_count",       "재시도횟수"),
    ("sources",           "데이터출처"),
    ("validation_reason", "비고"),
]


def save_to_excel(
    results: list[dict],
    manual_review: list[dict],
    output_dir: str = "output",
) -> str:
    """
    배당 데이터를 3시트 엑셀로 저장하고 파일 경로를 반환한다.
    """
    os.makedirs(output_dir, exist_ok=True)
    today = date.today().strftime("%Y%m%d")
    path = os.path.join(output_dir, f"dividend_result_{today}.xlsx")

    wb = openpyxl.Workbook()

    # 시트 1 — 배당 데이터
    ws1 = wb.active
    ws1.title = "배당 데이터"
    _write_sheet(ws1, results, _MAIN_COLS, header_color="1F4E79")

    # 시트 2 — 수동 확인 필요
    ws2 = wb.create_sheet("수동 확인 필요")
    _write_sheet(ws2, manual_review, _MANUAL_COLS, header_color="C00000")

    # 시트 3 — 검증 로그 (전체 합산)
    ws3 = wb.create_sheet("검증 로그")
    all_rows = [dict(r, validation_status="valid") for r in results] + \
               [dict(r, validation_status="manual_review") for r in manual_review]
    _write_sheet(ws3, all_rows, _LOG_COLS, header_color="375623")

    wb.save(path)
    logger.info("엑셀 저장 완료: %s  (valid=%d, manual=%d)", path, len(results), len(manual_review))
    return path


def _write_sheet(ws, rows: list[dict], col_defs: list[tuple], header_color: str):
    """시트에 헤더 + 데이터를 작성한다."""
    keys   = [c[0] for c in col_defs]
    labels = [c[1] for c in col_defs]

    # 헤더
    fill = PatternFill("solid", fgColor=header_color)
    font = Font(bold=True, color="FFFFFF")
    for col_idx, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    # 열 너비 자동 조정
    for col_idx, label in enumerate(labels, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(12, len(label) * 2)
